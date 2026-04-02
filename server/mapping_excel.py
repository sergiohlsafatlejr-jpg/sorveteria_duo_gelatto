#!/usr/bin/env python3
"""
mapping_excel.py — Utilitário para exportar/importar mapeamento PDV→Estoque via Excel.

Uso:
  python3 mapping_excel.py export <json_input_file> <output_xlsx>
  python3 mapping_excel.py import <input_xlsx> <output_json>
"""
import sys
import json
import os

def export_mapping(json_input_path: str, output_xlsx_path: str):
    """Gera um XLSX com os produtos do estoque e colunas editáveis para o código PDV."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    with open(json_input_path, "r", encoding="utf-8") as f:
        products = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapeamento PDV"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_blue = PatternFill("solid", fgColor="4F46E5")  # Produto estoque
    header_fill_green = PatternFill("solid", fgColor="059669")  # Colunas editáveis
    header_fill_gray = PatternFill("solid", fgColor="6B7280")   # Info
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    editable_fill = PatternFill("solid", fgColor="F0FDF4")  # Verde claro para editável
    mapped_fill = PatternFill("solid", fgColor="EFF6FF")    # Azul claro para já mapeado

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = "Mapeamento PDV → Estoque — Duo Gelatto"
    ws["A1"].font = Font(bold=True, size=14, color="1E1B4B")
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = "Preencha as colunas VERDES (Código PDV e Nome PDV) para vincular cada produto. Deixe em branco para remover o vínculo."
    ws["A2"].font = Font(size=9, color="6B7280", italic=True)
    ws["A2"].alignment = left

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # Cabeçalhos (linha 4)
    headers = [
        ("ID Produto", "A", 12, header_fill_blue),
        ("Nome no Estoque", "B", 35, header_fill_blue),
        ("Estoque Atual", "C", 14, header_fill_gray),
        ("Unidade", "D", 10, header_fill_gray),
        ("Código PDV ✏️", "E", 16, header_fill_green),
        ("Nome no PDV ✏️", "F", 30, header_fill_green),
        ("Status Atual", "G", 14, header_fill_gray),
    ]

    for i, (label, col, width, fill) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=label)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[4].height = 22

    # Dados
    for row_idx, prod in enumerate(products, 5):
        is_mapped = bool(prod.get("externalCode"))
        row_fill = mapped_fill if is_mapped else None

        # ID
        c = ws.cell(row=row_idx, column=1, value=prod["productId"])
        c.alignment = center
        c.border = border
        if row_fill: c.fill = row_fill

        # Nome estoque
        c = ws.cell(row=row_idx, column=2, value=prod["productName"])
        c.alignment = left
        c.border = border
        if row_fill: c.fill = row_fill

        # Estoque
        c = ws.cell(row=row_idx, column=3, value=prod.get("currentStock", 0))
        c.alignment = center
        c.border = border
        if row_fill: c.fill = row_fill

        # Unidade
        c = ws.cell(row=row_idx, column=4, value=prod.get("unit", "un"))
        c.alignment = center
        c.border = border
        if row_fill: c.fill = row_fill

        # Código PDV (editável)
        ext_code = prod.get("externalCode") or ""
        c = ws.cell(row=row_idx, column=5, value=ext_code)
        c.alignment = center
        c.border = border
        c.fill = editable_fill

        # Nome PDV (editável)
        ext_name = prod.get("externalName") or ""
        c = ws.cell(row=row_idx, column=6, value=ext_name)
        c.alignment = left
        c.border = border
        c.fill = editable_fill

        # Status
        status = "✅ Mapeado" if is_mapped else "⚠️ Sem vínculo"
        c = ws.cell(row=row_idx, column=7, value=status)
        c.alignment = center
        c.border = border
        if row_fill: c.fill = row_fill

        ws.row_dimensions[row_idx].height = 18

    # Congelar cabeçalho
    ws.freeze_panes = "A5"

    # Legenda
    last_row = len(products) + 6
    ws.cell(row=last_row, column=1, value="Legenda:")
    ws.cell(row=last_row, column=1).font = Font(bold=True)
    ws.cell(row=last_row + 1, column=1, value="• Colunas com fundo VERDE são editáveis")
    ws.cell(row=last_row + 2, column=1, value="• Preencha 'Código PDV' com o código numérico do PDV (ex: 146)")
    ws.cell(row=last_row + 3, column=1, value="• 'Nome no PDV' é opcional — apenas para referência")
    ws.cell(row=last_row + 4, column=1, value="• Deixe 'Código PDV' em branco para remover o vínculo existente")
    ws.cell(row=last_row + 5, column=1, value="• Salve o arquivo e importe de volta no sistema")
    for r in range(last_row, last_row + 6):
        ws.cell(row=r, column=1).font = Font(size=9, color="6B7280", italic=True)

    wb.save(output_xlsx_path)
    print(json.dumps({"success": True, "rows": len(products), "file": output_xlsx_path}))


def import_mapping(input_xlsx_path: str, output_json_path: str):
    """Lê o XLSX preenchido e retorna JSON com os vínculos para salvar no banco."""
    from openpyxl import load_workbook

    wb = load_workbook(input_xlsx_path, data_only=True)
    ws = wb.active

    results = []
    errors = []

    # Dados começam na linha 5 (linha 4 = cabeçalho)
    for row in ws.iter_rows(min_row=5, values_only=True):
        product_id = row[0]  # Coluna A
        product_name = row[1]  # Coluna B
        external_code_raw = row[4]  # Coluna E — Código PDV
        external_name = row[5]  # Coluna F — Nome PDV (opcional)

        # Pular linhas vazias ou de legenda
        if product_id is None or not str(product_id).strip():
            continue
        try:
            pid = int(float(str(product_id).strip()))
        except (ValueError, TypeError):
            continue

        # Código PDV: converter para string limpa ou None
        ext_code = None
        if external_code_raw is not None:
            raw = str(external_code_raw).strip()
            if raw and raw.lower() not in ("", "none", "nan"):
                # Remover .0 de números
                if raw.endswith(".0"):
                    raw = raw[:-2]
                ext_code = raw

        results.append({
            "productId": pid,
            "productName": str(product_name or "").strip(),
            "externalCode": ext_code,
            "externalName": str(external_name or "").strip() if external_name else None,
        })

    output = {
        "success": True,
        "total": len(results),
        "toLink": len([r for r in results if r["externalCode"]]),
        "toUnlink": len([r for r in results if not r["externalCode"]]),
        "mappings": results,
    }

    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False)

    print(json.dumps({"success": True, "total": len(results)}))


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print(json.dumps({"error": "Uso: mapping_excel.py [export|import] <input> <output>"}))
        sys.exit(1)

    mode = sys.argv[1]
    inp = sys.argv[2]
    out = sys.argv[3]

    if mode == "export":
        export_mapping(inp, out)
    elif mode == "import":
        import_mapping(inp, out)
    else:
        print(json.dumps({"error": f"Modo inválido: {mode}"}))
        sys.exit(1)
